# Modifica las importaciones
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from tkinterdnd2 import DND_FILES
from PIL import Image, ImageTk
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from database.queries import fetch_student_by_rut, insert_student, update_student, enroll_student
from path_utils import resource_path

class BulkEnrollment:
    def __init__(self, parent):
        self.window = tk.Toplevel(parent)
        self.window.title("Matrícula e Inscripción Masiva")
        self.window.grab_set()
        
        # Centrar y dimensionar la ventana
        window_width = 600
        window_height = 500
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        self.window.geometry(f'{window_width}x{window_height}+{x}+{y}')
        
        try:
            self.window.iconbitmap(resource_path('assets/logo2.ico'))
        except Exception as e:
            print(f"Error al cargar ícono: {e}")
        
        self.window.configure(bg='#f0f0f0')
        self.setup_ui()

    def setup_styles(self):
        style = ttk.Style()
        style.configure('Action.TButton',
                font=('Helvetica', 10, 'bold'),
                padding=(10, 5),
                background='#00239c',
                foreground='white',
                relief='raised',  # Cambiado a raised para dar el efecto 3D
                borderwidth=1)    # Añadido borde

        style.map('Action.TButton',
                    background=[('active', '#001970'),
                            ('pressed', '#00239c')],
                    foreground=[('active', 'white'),
                            ('pressed', 'white')],
                    relief=[('pressed', 'sunken')])  # Efecto presionado

        style.configure('delete.TButton',
                        font=('Helvetica', 10, 'bold'),
                        padding=(10, 5),
                        background='#b50707',
                        foreground='white',
                        relief='raised',  # Cambiado a raised para dar el efecto 3D
                        borderwidth=1)    # Añadido borde

        style.map('delete.TButton',
                    background=[('active', '#990606'),
                            ('pressed', '#b50707')],
                    foreground=[('active', 'white'),
                            ('pressed', 'white')],
                    relief=[('pressed', 'sunken')])  # Efecto presionado

    def setup_ui(self):
        # Frame principal
        main_frame = tk.Frame(self.window, bg='#f0f0f0')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        # Título - usando tk.Label en lugar de ttk.Label
        title_label = tk.Label(
            main_frame,
            text="Matrícula e Inscripción Masiva",
            font=('Segoe UI', 16, 'bold'),
            foreground='#00239c',
            bg='#f0f0f0'
        )
        title_label.pack(pady=(0, 20))

        # Frame para los pasos con fondo gris
        steps_frame = tk.Frame(main_frame, bg='#e8e8e8')
        steps_frame.pack(fill=tk.X, pady=(0, 20))

        # Paso 1
        step1_frame = tk.Frame(steps_frame, bg='#e8e8e8')
        step1_frame.pack(fill=tk.X, pady=5, padx=10)
        tk.Label(
            step1_frame,
            text="1. Descarga la plantilla y complétala:",
            font=('Segoe UI', 10),
            bg='#e8e8e8',
            anchor='w'
        ).pack(side=tk.LEFT)
        ttk.Button(
            step1_frame,
            text="Descargar Plantilla",
            style='Action.TButton',
            command=self.download_template
        ).pack(side=tk.RIGHT)

        # Paso 2
        step2_frame = tk.Frame(steps_frame, bg='#e8e8e8')
        step2_frame.pack(fill=tk.X, pady=5, padx=10)
        tk.Label(
            step2_frame,
            text="2. Carga el Excel completado:",
            font=('Segoe UI', 10),
            bg='#e8e8e8',
            anchor='w'
        ).pack(side=tk.LEFT)
        ttk.Button(
            step2_frame,
            text="Cargar Excel",
            style='Action.TButton',
            command=self.upload_excel
        ).pack(side=tk.RIGHT)

        # Área de drop
        self.drop_frame = tk.Frame(
            main_frame,
            bg='white',
            highlightthickness=1,
            highlightbackground='#cccccc'
        )
        self.drop_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        # Contenido del área de drop
        drop_content = tk.Frame(self.drop_frame, bg='white')
        drop_content.place(relx=0.5, rely=0.5, anchor="center")

        # Icono de documento
        doc_label = tk.Label(
            drop_content,
            text="📄",
            font=('Segoe UI', 48),
            bg='white'
        )
        doc_label.pack()

        # Texto de instrucción
        instruction_label = tk.Label(
            drop_content,
            text="Arrastra aquí tu archivo Excel",
            font=('Segoe UI', 12),
            bg='white',
            fg='#666'
        )
        instruction_label.pack(pady=(10, 5))

        separator_label = tk.Label(
            drop_content,
            text="o",
            font=('Segoe UI', 10),
            bg='white',
            fg='#666'
        )
        separator_label.pack(pady=2)

        # Botón en área de drop
        select_button = ttk.Button(
            drop_content,
            text="Seleccionar Archivo",
            style='Action.TButton',
            command=self.upload_excel
        )
        select_button.pack(pady=(5, 0))

        # Configurar el drop
        try:
            self.drop_frame.drop_target_register(DND_FILES)
            self.drop_frame.dnd_bind('<<Drop>>', self.handle_drop)
        except Exception as e:
            print(f"Error configurando DnD: {e}")

        # Frame para el preview (inicialmente oculto)
        self.preview_frame = tk.Frame(main_frame, bg='#f0f0f0')
        self.preview_frame.pack_forget()

        # Configurar Treeview
        self.tree = ttk.Treeview(
            self.preview_frame,
            show='headings'
        )
        
        # Scrollbars para el Treeview
        self.scrolly = ttk.Scrollbar(self.preview_frame, orient="vertical", command=self.tree.yview)
        self.scrollx = ttk.Scrollbar(self.preview_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=self.scrolly.set, xscrollcommand=self.scrollx.set)

        # Botón de procesar (inicialmente oculto)
        self.process_button = ttk.Button(
            main_frame,
            text="Procesar Matrículas",
            style='Action.TButton',
            command=self.process_enrollments
        )

    def show_preview(self, df):
        """Muestra la previsualización de datos"""
        # Ocultar área de drop
        self.drop_frame.pack_forget()
        
        # Mostrar Treeview
        self.tree_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.scrolly.pack(side=tk.RIGHT, fill=tk.Y)
        self.scrollx.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Limpiar y configurar columnas
        for col in self.tree['columns']:
            self.tree.heading(col, text='')
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        # Configurar nuevas columnas
        self.tree['columns'] = list(df.columns)
        for col in df.columns:
            self.tree.heading(col, text=col)
            # Ajustar ancho basado en contenido
            max_width = max(
                len(str(col)),
                df[col].astype(str).str.len().max()
            ) * 10
            self.tree.column(col, width=min(max_width, 300))
            
        # Insertar datos
        for idx, row in df.iterrows():
            self.tree.insert('', 'end', values=list(row))
            
        # Mostrar botón de proceso
        self.process_button.pack(pady=10)

    def handle_drop(self, event):
        """Maneja el evento de soltar archivos"""
        file_path = event.data
        if isinstance(file_path, str):
            # Limpiar la ruta (eliminar {} en Windows)
            file_path = file_path.strip('{}')
            
            if file_path.lower().endswith('.xlsx'):
                self.process_excel_file(file_path)
            else:
                messagebox.showwarning(
                    "Formato Incorrecto",
                    "Por favor, selecciona un archivo Excel (.xlsx)"
                )

    def process_excel_file(self, file_path):
        """Procesa el archivo Excel seleccionado"""
        try:
            df = pd.read_excel(file_path)
            
            # Ocultar área de drop
            self.drop_frame.pack_forget()
            
            # Mostrar y configurar treeview
            self.preview_frame.pack(fill=tk.BOTH, expand=True, pady=10)
            self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            self.scrolly.pack(side=tk.RIGHT, fill=tk.Y)
            self.scrollx.pack(side=tk.BOTTOM, fill=tk.X)
            
            # Limpiar treeview anterior si existe
            for col in self.tree['columns']:
                self.tree.heading(col, text='')
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            # Configurar columnas
            self.tree['columns'] = list(df.columns)
            for col in df.columns:
                self.tree.heading(col, text=col)
                # Ajustar ancho basado en contenido
                max_width = max(
                    len(str(col)),
                    df[col].astype(str).str.len().max()
                ) * 10
                self.tree.column(col, width=min(max_width, 300))
            
            # Insertar datos
            for idx, row in df.iterrows():
                self.tree.insert('', 'end', values=list(row))
            
            # Mostrar botón de proceso
            self.process_button.pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al procesar el archivo: {str(e)}")

    def upload_excel(self):
        """Maneja la selección de archivo mediante diálogo"""
        file_path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Archivos Excel", "*.xlsx")]
        )
        if file_path:
            self.process_excel_file(file_path)

    def download_template(self):
        """Genera y descarga la plantilla Excel con estilos"""
        try:
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            wb = Workbook()
            ws = wb.active
            ws.title = "Matricula Masiva"
            
            # Definir estilos
            header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            
            # Colores para las diferentes secciones
            alumno_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # Verde claro
            curso_fill = PatternFill(start_color="CCEEFF", end_color="CCEEFF", fill_type="solid")   # Azul claro
            pago_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")    # Rojo claro
            
            # Borde para las celdas
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Encabezados por sección
            headers = {
                "Datos del Alumno": [
                    "RUT*", "Nombre*", "Apellido*", "Email*", "Teléfono*", 
                    "Profesión", "Dirección", "Ciudad", "Comuna"
                ],
                "Datos del Curso": [
                    "ID Curso*", "N° Acta", "Fecha Inscripción*", "Fecha Término",
                    "Año*", "Método*"
                ],
                "Datos de Pago": [
                    "Empresa", "Código SENCE", "Folio"
                ]
            }

            # Escribir los encabezados con estilos
            col = 1
            for section, fields in headers.items():
                for field in fields:
                    cell = ws.cell(row=1, column=col, value=field)
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Aplicar color según la sección
                    if section == "Datos del Alumno":
                        cell.fill = alumno_fill
                    elif section == "Datos del Curso":
                        cell.fill = curso_fill
                    else:
                        cell.fill = pago_fill
                    
                    col += 1

            # Datos de ejemplo
            example_data = [
                "12345678-9", "Juan", "Pérez", "juan@email.com", "+56912345678",
                "Ingeniero", "Calle 123", "Santiago", "Las Condes",
                "CURSO001", "ACTA001", "2024-01-15", "2024-02-15",
                "2024", "Presencial", "Empresa SA", "SENCE001", "FOL001"
            ]
            
            # Insertar datos de ejemplo
            for col, value in enumerate(example_data, 1):
                cell = ws.cell(row=2, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Ajustar anchos de columna
            for column_cells in ws.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = max(length + 2, 15)

            # Altura de filas
            ws.row_dimensions[1].height = 30  # Altura para encabezados
            ws.row_dimensions[2].height = 25  # Altura para ejemplo

            # Agregar una nota sobre los campos requeridos
            ws['A4'] = "* Campos obligatorios"
            ws['A4'].font = Font(italic=True)

            # Guardar archivo
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile="plantilla_matricula.xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if file_path:
                wb.save(file_path)
                messagebox.showinfo(
                    "Éxito", 
                    "Plantilla descargada correctamente\n\n" +
                    "Verde: Datos del Alumno\n" +
                    "Azul: Datos del Curso\n" +
                    "Rojo: Datos de Pago\n\n" +
                    "Los campos con * son obligatorios"
                )
                
        except Exception as e:
            messagebox.showerror("Error", f"Error al crear plantilla: {str(e)}")

    def process_enrollments(self):
        """Procesa las matrículas masivas"""
        try:
            if not self.tree.get_children():
                messagebox.showwarning("Advertencia", "No hay datos para procesar")
                return

            success_count = 0
            error_count = 0
            error_messages = []

            for item in self.tree.get_children():
                try:
                    values = self.tree.item(item)['values']
                    # Procesar cada registro
                    student_data = {
                        'rut': values[0],
                        'nombre': values[1],
                        'apellido': values[2],
                        'email': values[3],
                        'telefono': values[4],
                        'profesion': values[5],
                        'direccion': values[6],
                        'ciudad': values[7],
                        'comuna': values[8]
                    }

                    # Verificar si existe el alumno
                    existing_student = fetch_student_by_rut(values[0])
                    if existing_student:
                        update_student(student_data)
                    else:
                        insert_student(student_data)

                    # Crear inscripción
                    inscription_data = {
                        'rut_alumno': values[0],
                        'id_curso': values[9],
                        'num_acta': values[10],
                        'fecha_inscripcion': values[11],
                        'fecha_termino': values[12],
                        'anio': values[13],
                        'metodo': values[14],
                        'cod_sence': values[15],
                        'folio': values[16]
                    }
                    
                    enroll_student(inscription_data)
                    success_count += 1

                except Exception as e:
                    error_count += 1
                    error_messages.append(f"Error en línea {success_count + error_count}: {str(e)}")

            # Mostrar resumen
            message = f"Proceso completado:\n\n"
            message += f"Registros exitosos: {success_count}\n"
            if error_count > 0:
                message += f"Registros con error: {error_count}\n\n"
                message += "Detalle de errores:\n" + "\n".join(error_messages)

            if error_count > 0:
                messagebox.showwarning("Proceso Completado con Advertencias", message)
            else:
                messagebox.showinfo("Proceso Completado", message)
                self.window.destroy()

        except Exception as e:
            messagebox.showerror("Error", f"Error en el proceso: {str(e)}")